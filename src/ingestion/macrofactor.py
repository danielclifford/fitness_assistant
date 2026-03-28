"""
MacroFactor Excel importer.

How to export:
  MacroFactor app → Profile → Export Data → Export

The export is an .xlsx file with multiple sheets:
  - Calories & Macros   : Date, Calories, Fat, Carbs, Protein
  - Micronutrients      : Date, Fiber (g), and many others
  - Scale Weight        : Date, Weight (kg), Fat Percent
  - Weight Trend        : Date, Trend Weight (kg)  [MacroFactor's smoothed weight]
  - Expenditure         : Date, Expenditure  [MacroFactor's daily TDEE estimate]
  - Body Metrics        : Date, body measurements (waist, bicep, etc.)

We import from: Calories & Macros, Micronutrients (fiber only), Scale Weight,
Expenditure. Weight Trend is also stored as it's MacroFactor's smoothed signal.
"""

from datetime import date, datetime
from pathlib import Path
import openpyxl
from src.database import get_connection


def _to_date(val) -> date | None:
    """Convert openpyxl cell value to a date object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _sheet_by_keywords(wb, *keywords) -> openpyxl.worksheet.worksheet.Worksheet | None:
    """Find a sheet whose name contains all keywords (case-insensitive)."""
    for name in wb.sheetnames:
        lower = name.lower()
        if all(k.lower() in lower for k in keywords):
            return wb[name]
    return None


def _read_sheet(sheet, date_col: int = 0) -> dict[str, dict]:
    """
    Read a sheet into {date_str: {col_header: value}}.
    date_col is the 0-based index of the date column.
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = {}
    for row in rows[1:]:
        if not row or row[date_col] is None:
            continue
        d = _to_date(row[date_col])
        if d is None:
            continue
        result[d.isoformat()] = {
            headers[i]: row[i]
            for i in range(len(row))
            if i < len(headers) and i != date_col
        }
    return result


def import_xlsx(filepath: str | Path) -> dict:
    """
    Parse a MacroFactor .xlsx export and upsert into nutrition_logs and weight_logs.
    Returns counts of rows written to each table.
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # --- Load each relevant sheet ---
    macros_sheet      = _sheet_by_keywords(wb, "calorie") or _sheet_by_keywords(wb, "macro")
    micros_sheet      = _sheet_by_keywords(wb, "micronutrient")
    weight_sheet      = _sheet_by_keywords(wb, "scale", "weight") or _sheet_by_keywords(wb, "weight")
    trend_sheet       = _sheet_by_keywords(wb, "trend")
    expenditure_sheet = _sheet_by_keywords(wb, "expenditure")

    macros_data      = _read_sheet(macros_sheet)      if macros_sheet      else {}
    micros_data      = _read_sheet(micros_sheet)      if micros_sheet      else {}
    weight_data      = _read_sheet(weight_sheet)      if weight_sheet      else {}
    trend_data       = _read_sheet(trend_sheet)       if trend_sheet       else {}
    expenditure_data = _read_sheet(expenditure_sheet) if expenditure_sheet else {}

    wb.close()

    # --- Find the fiber column header (varies by locale/version) ---
    fiber_col = None
    if micros_data:
        sample = next(iter(micros_data.values()))
        for k in sample:
            if "fiber" in k.lower() or "fibre" in k.lower():
                fiber_col = k
                break

    conn = get_connection()
    nutrition_count = 0
    weight_count = 0

    # Collect all dates across nutrition and expenditure sheets
    all_nutrition_dates = set(macros_data) | set(expenditure_data)

    for d in sorted(all_nutrition_dates):
        macros = macros_data.get(d, {})
        micros = micros_data.get(d, {})
        expend = expenditure_data.get(d, {})

        calories  = _to_float(macros.get("Calories (kcal)"))
        fat_g     = _to_float(macros.get("Fat (g)"))
        carbs_g   = _to_float(macros.get("Carbs (g)"))
        protein_g = _to_float(macros.get("Protein (g)"))
        fiber_g   = _to_float(micros.get(fiber_col)) if fiber_col else None
        mf_tdee   = _to_float(next(iter(expend.values()), None)) if expend else None

        # A day with no macros recorded but an expenditure entry = unlogged day
        # A day with very low calories is also flagged as incomplete
        is_complete = 1
        if calories is None and protein_g is None:
            is_complete = 0
        elif calories is not None and calories < 400:
            is_complete = 0

        if calories is not None or mf_tdee is not None:
            conn.execute("""
                INSERT INTO nutrition_logs
                    (date, calories, protein_g, carbs_g, fat_g, fiber_g,
                     macrofactor_tdee, is_complete_log, source)
                VALUES (?,?,?,?,?,?,?,?,'macrofactor')
                ON CONFLICT(date) DO UPDATE SET
                    calories          = excluded.calories,
                    protein_g         = excluded.protein_g,
                    carbs_g           = excluded.carbs_g,
                    fat_g             = excluded.fat_g,
                    fiber_g           = excluded.fiber_g,
                    macrofactor_tdee  = excluded.macrofactor_tdee,
                    is_complete_log   = excluded.is_complete_log
            """, (d, calories, protein_g, carbs_g, fat_g, fiber_g, mf_tdee, is_complete))
            nutrition_count += 1

    # --- Weight entries from Scale Weight sheet ---
    for d, vals in sorted(weight_data.items()):
        weight_kg = _to_float(vals.get("Weight (kg)"))
        if weight_kg is not None:
            conn.execute("""
                INSERT INTO weight_logs (date, weight_kg, source)
                VALUES (?, ?, 'macrofactor')
                ON CONFLICT(date, source) DO UPDATE SET weight_kg = excluded.weight_kg
            """, (d, weight_kg))
            weight_count += 1

    # --- Trend weight (MacroFactor's smoothed signal) ---
    # Store alongside raw weight so TDEE analysis can use it
    for d, vals in sorted(trend_data.items()):
        trend_kg = _to_float(vals.get("Trend Weight (kg)"))
        if trend_kg is not None:
            conn.execute("""
                INSERT INTO weight_logs (date, weight_kg, source)
                VALUES (?, ?, 'macrofactor_trend')
                ON CONFLICT(date, source) DO UPDATE SET weight_kg = excluded.weight_kg
            """, (d, trend_kg))

    conn.commit()
    conn.close()

    sheets_found = []
    if macros_sheet:      sheets_found.append("Calories & Macros")
    if micros_sheet:      sheets_found.append("Micronutrients")
    if weight_sheet:      sheets_found.append("Scale Weight")
    if expenditure_sheet: sheets_found.append("Expenditure")

    return {
        "nutrition_rows": nutrition_count,
        "weight_rows":    weight_count,
        "sheets_imported": sheets_found,
    }
